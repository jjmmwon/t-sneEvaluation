import glob
import os
import argparse
import json

import pandas as pd
import numpy as np

from openpyxl import Workbook, load_workbook



def argparsing():
    parser = argparse.ArgumentParser(description="Rate the PCA")
    parser.add_argument('--file_path', '-f', default='default', help="File path to convert json to csv")

    args = parser.parse_args()

    return args

def main():
    args = argparsing()

    save_path = "/home/myeongwon/mw/tsne/result/"
    excel_path = save_path + "result.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    wb.save(excel_path)

    pca_rank_cnt =[ [0 for _ in range(11)] for _ in range(3) ] 
    pca_rank_sum = []
    rand_rank_sum = []
    better_than_pca = []
    better_than_pca_cnt = [0 for _ in range(11)]

    if args.file_path == 'default':
        path = '/home/myeongwon/mw/tsne/result/*/*.json'
        json_files = glob.glob(path)
        json_files.remove('/home/myeongwon/mw/tsne/result/HepatitisCdata/HepatitisCdata_result.json')
        

        r = 1


        for json_file in json_files:
            with open(json_file,'r') as f:
                json_data = json.load(f)
            print(json_file)
            for data in json_data:
                pca = data['pca'][0]
                rand = data['random']

                if len(data['random'])  != 10 :
                    print(json_file)
                    continue

                pca_data = ['PCA', round(pca['Loss'], 5), round(pca['Steadiness'], 5), round(pca['Cohesiveness'], 5) ]
                rand_data = [ [ f'Random{i}', round(LSC['Loss'], 5), round(LSC['Steadiness'], 5), round(LSC["Cohesiveness"], 5) ] for i, LSC in enumerate(rand, start=1)]

                l_list = []
                s_list = []
                c_list = []

                for LSC in rand:
                    l_list.append(round(LSC['Loss'], 5))
                    s_list.append(round(LSC['Steadiness'], 5))
                    c_list.append(round(LSC['Cohesiveness'], 5))

                rand_mean = [np.mean(l_list), np.mean(s_list), np.mean(c_list)]

                l_list.append(round(pca['Loss'], 5))
                s_list.append(round(pca['Steadiness'], 5))
                c_list.append(round(pca['Cohesiveness'], 5))

                l_list.sort()
                s_list.sort(reverse=True)
                c_list.sort(reverse=True)

                pca_rank = [ l_list.index(round(pca['Loss'], 5))+1, s_list.index(round(pca['Steadiness'], 5))+1, c_list.index(round(pca['Cohesiveness'], 5))+1 ]
                pca_rank_cnt[0][l_list.index(round(pca['Loss'], 5))] += 1
                pca_rank_cnt[1][s_list.index(round(pca['Steadiness'], 5))] += 1
                pca_rank_cnt[2][c_list.index(round(pca['Cohesiveness'], 5))] += 1
                pca_rank_sum.append( l_list.index(round(pca['Loss'], 5))+1 + s_list.index(round(pca['Steadiness'], 5))+1 + c_list.index(round(pca['Cohesiveness'], 5))+1 )

                for LSC in rand:
                    rand_rank_sum.append( l_list.index(round(LSC['Loss'], 5))+1 + s_list.index(round(LSC['Steadiness'], 5))+1 + c_list.index(round(LSC['Cohesiveness'], 5))+1 )

                cnt = 0
                for i in range(-1,-11,-1):
                    if pca_rank_sum[-1] > rand_rank_sum[i]:
                        cnt += 1
                better_than_pca.append(cnt)
                better_than_pca_cnt[cnt] += 1

                data_dict = {}
                for rd in rand_data:
                    data_dict[rd[0]] = rd[1:]
                data_dict['Random_Mean'] = rand_mean
                data_dict['PCA'] = pca_data[1:]
                data_dict['PCA_Rank'] = pca_rank

                df = pd.DataFrame(data_dict, index=['Loss', 'Steadiness', 'Cohesiveness'])
                header = {}
                header['Dataset'] = json_file[json_file.rfind('/')+1:json_file.rfind('.')]
                header['Shape'] = f"Instances: { data['Shape']['Instances'] }, Attributes: { data['Shape']['Attributes']}"
                header['Hyperparameter'] = f"Perplexity: {data['Hyperparameter']['perplexity']}, Iterations: {data['Hyperparameter']['max_iter']}, Learning Rate: {data['Hyperparameter']['learning_rate']}"

                wb = load_workbook(excel_path)
                ws = wb['Sheet1']
                ws.cell(row=r, column=1, value='Dataset');   ws.cell(row=r, column=2, value=header['Dataset'])
                ws.cell(row=r, column=3, value='Shape'); ws.cell(row=r, column=4, value= header['Shape'])
                ws.cell(row=r, column=5, value='Hyperparameter');   ws.cell(row=r, column=6, value= header['Hyperparameter'])

                wb.save(excel_path)
                


                with pd.ExcelWriter(excel_path, mode='a', if_sheet_exists = 'overlay') as writer:  
                    df.to_excel(writer, startcol=1, startrow=r, engine='openpyxl')

                r = r + 6

        pca_rank_dict= {}
        pca_rank_dict['Loss_Rank_Count'] = pca_rank_cnt[0]
        pca_rank_dict['Steadomess_Rank_Count'] = pca_rank_cnt[1]
        pca_rank_dict['Cohesiveness_Rank_Count'] = pca_rank_cnt[2]
        pca_rank_df = pd.DataFrame(pca_rank_dict, index=[i for i in range(1,12)])

        better_than_pca_cnt_df = pd.DataFrame({'Better_than_PCA_cnt' : better_than_pca_cnt}, index=[i for i in range(11)])

        with pd.ExcelWriter(excel_path, mode='a', if_sheet_exists = 'overlay') as writer:  
            pca_rank_df.to_excel(writer, startcol= 16, startrow=1, engine='openpyxl')
            better_than_pca_cnt_df.to_excel(writer, startcol= 16, startrow=17, engine='openpyxl')




        import seaborn as sns
        import matplotlib.pyplot as plt

        title = ['Loss', 'Steadiness', 'Cohesiveness']
        for i in range(3):
            plt.plot( list(range(1,12)), pca_rank_cnt[i])
            plt.title(f'PCA {title[i]} Rank Distribution')
            plt.savefig(save_path + f'PCA_{title[i]}_rank_dist.png')
            plt.cla()
        
        sns.histplot(x=better_than_pca)
        plt.title(f'Distribution of the number of random better than pca')
        plt.savefig(save_path + f'better_than_PCA_dist.png')
        plt.cla()

        sns.kdeplot(x=pca_rank_sum )
        sns.kdeplot(x=rand_rank_sum, color='Red')
        plt.legend(('PCA', 'Random'))
        plt.title(f'Each Rank Sum Distribution')
        plt.savefig(save_path + f'each_rank_sum_dist.png')
        plt.cla()

        sns.kdeplot(x=rand_rank_sum)
        plt.title(f'Random Rank Sum Distribution')
        plt.savefig(save_path + f'Random_rank_sum_dist.png')

    else:
        pass

if __name__ == '__main__':
    main()